"""Leitura de planilhas (CSV e XLSX) com tratamento de encoding e decimal.

A planilha do Canvas vem com algumas pegadinhas:
  - encoding Latin-1 em CSV (acentos viram "Ã­", "Â°")
  - vírgula como separador decimal ("7,85")
  - segunda linha é "Points Possible" (nota máxima por coluna), não um aluno
  - células vazias significam falta, NÃO zero

Aqui só lemos e tipamos. Validação semântica fica no header.py / pipeline.py.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass


@dataclass(frozen=True)
class PlanilhaCrua:
    """Resultado do parsing bruto do arquivo, antes de qualquer interpretação."""
    cabecalho: list[str]
    pontos_possiveis: list[str | None]
    linhas: list[list[str | None]]   # mesma largura do cabecalho; None = célula vazia


# ─── Detecção de encoding ─────────────────────────────────────────────────


ENCODINGS_TENTADOS = ("utf-8-sig", "utf-8", "latin-1", "cp1252")


def decodificar(bytes_: bytes) -> str:
    """Tenta múltiplos encodings comuns. Lança ValueError se nenhum funcionar."""
    for enc in ENCODINGS_TENTADOS:
        try:
            return bytes_.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError(
        f"Não foi possível decodificar o arquivo com nenhum dos encodings {ENCODINGS_TENTADOS}"
    )


# ─── CSV ──────────────────────────────────────────────────────────────────


def ler_csv(conteudo: bytes) -> PlanilhaCrua:
    """Parseia bytes de um CSV exportado do Canvas.

    O Canvas exporta em dois formatos no Brasil:
      - separador `,` com decimais quoted (`"7,85"`)
      - separador `;` com decimais raw (`7,85`)
    Detectamos automaticamente qual está em uso.
    """
    texto = decodificar(conteudo)
    delimitador = _detectar_delimitador(texto)
    leitor = csv.reader(io.StringIO(texto), delimiter=delimitador)
    linhas_brutas = [linha for linha in leitor]

    if len(linhas_brutas) < 2:
        raise ValueError("CSV precisa ter pelo menos cabeçalho e linha 'Points Possible'.")

    cabecalho = [c.strip() for c in linhas_brutas[0]]
    pontos = _normalizar_linha(linhas_brutas[1], largura=len(cabecalho))

    if (pontos[0] or "").strip().lower() != "points possible":
        raise ValueError(
            "Segunda linha do CSV deve ser 'Points Possible' (export padrão do Canvas). "
            f"Encontrei: {pontos[0]!r}."
        )

    linhas_alunos = [
        _normalizar_linha(linha, largura=len(cabecalho))
        for linha in linhas_brutas[2:]
        if any((c or "").strip() for c in linha)  # descarta linhas totalmente vazias
    ]

    return PlanilhaCrua(
        cabecalho=cabecalho,
        pontos_possiveis=pontos,
        linhas=linhas_alunos,
    )


def _detectar_delimitador(texto: str) -> str:
    """Decide entre ',' e ';' analisando a primeira linha.

    Heurística: o delimitador certo aparece muito mais vezes que o errado
    na linha de cabeçalho (porque o cabeçalho do Canvas tem ~50+ colunas).
    Em empate ou indecisão, prioriza ';' (export PT-BR mais comum).
    """
    primeira_linha = texto.split("\n", 1)[0]
    n_virgulas = primeira_linha.count(",")
    n_pontos_virgulas = primeira_linha.count(";")
    if n_pontos_virgulas >= n_virgulas:
        return ";"
    return ","


def _normalizar_linha(linha: list[str], largura: int) -> list[str | None]:
    """Preenche/trunca para largura fixa e converte string vazia em None."""
    normalizada: list[str | None] = []
    for i in range(largura):
        valor = linha[i] if i < len(linha) else ""
        valor = (valor or "").strip()
        normalizada.append(valor if valor != "" else None)
    return normalizada


# ─── XLSX (opcional, ativa quando openpyxl está instalado) ────────────────


def ler_xlsx(conteudo: bytes) -> PlanilhaCrua:
    """Parseia um XLSX (primeira aba). Requer openpyxl no ambiente."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Suporte a XLSX requer o pacote 'openpyxl' instalado. "
            "Adicione ao requirements.txt e reinstale."
        ) from exc

    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("XLSX sem abas legíveis.")

    linhas_brutas: list[list[str]] = []
    for linha in ws.iter_rows(values_only=True):
        linhas_brutas.append(["" if c is None else str(c) for c in linha])

    if len(linhas_brutas) < 2:
        raise ValueError("XLSX precisa ter pelo menos cabeçalho e linha 'Points Possible'.")

    cabecalho = [c.strip() for c in linhas_brutas[0]]
    pontos = _normalizar_linha(linhas_brutas[1], largura=len(cabecalho))

    if (pontos[0] or "").strip().lower() != "points possible":
        raise ValueError(
            "Segunda linha do XLSX deve ser 'Points Possible' (export padrão do Canvas)."
        )

    linhas_alunos = [
        _normalizar_linha(linha, largura=len(cabecalho))
        for linha in linhas_brutas[2:]
        if any((c or "").strip() for c in linha)
    ]

    return PlanilhaCrua(
        cabecalho=cabecalho,
        pontos_possiveis=pontos,
        linhas=linhas_alunos,
    )


# ─── Helpers numéricos ────────────────────────────────────────────────────


def parsear_decimal_br(valor: str | None) -> float | None:
    """'7,85' → 7.85. '' / None → None. Aceita '7.85' também."""
    if valor is None:
        return None
    s = valor.strip().replace(",", ".")
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None
